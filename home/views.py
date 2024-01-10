from django.http import StreamingHttpResponse
from django.shortcuts import redirect, render, reverse
import mimetypes
import os
from django.http.response import HttpResponse
from wsgiref.util import FileWrapper

import openpyxl
from home.models import Item, Update

from home.authentication import new_session
from home.fetching import fetching_manager
from home.updating import update_database


def index(request):
    return render(request, 'home/homepage.html') 

def all_items(request):
    u = Update.objects.first()
    if u is None:
        # set up default
        return render(request, 'home/all_items.html', {}) 
    else:
        items = Item.objects.order_by('sku')
        return render(request, 'home/all_items.html', {"items": items, "date": u.updated_at})

def out_stock(request):
    u = Update.objects.first()
    if u is None:
        # set up default
        return render(request, 'home/outstock.html', {}) 
    else:
        items = Item.objects.filter(available__lte=0).order_by('sku')
        return render(request, 'home/outstock.html', {"items": items, "date": u.updated_at}) 
    
def incoming(request):
    u = Update.objects.first()
    if u is None:
        # set up default
        return render(request, 'home/incoming.html', {}) 
    else:
        items = Item.objects.filter(incoming__gt=0).order_by('sku')
        return render(request, 'home/incoming.html', {"items": items, "date": u.updated_at})

def update_page(request, page=""):
    page = request.GET.get('page')
    
    # start session
    client = new_session()
        
    items = fetching_manager(client)
        
    update_database(items)
        
    items = Item.objects.all()
    
    if page == "all":
        return redirect(all_items)
    elif page == "incoming":
        return redirect(incoming) 
    elif page == "out":
        return redirect(out_stock)
    else:   
        return redirect(index) 

  
def export_excel(request):
    u = Update.objects.first()
    if u is not None:

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="inventoryExport.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Inventory'

        # Write header row
        header = ['SKU', 'Available', 'Incoming']
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        prev = request.META.get('HTTP_REFERER')
        if "incoming" in prev:
            queryset = Item.objects.order_by('sku').filter(incoming__gt=0).values_list('sku', 'available','incoming')
        elif "out_of_stock" in prev:
            queryset = Item.objects.order_by('sku').filter(available__lte=0).values_list('sku', 'available','incoming') 
        else: 
            queryset = Item.objects.order_by('sku').values_list('sku', 'available','incoming') 
        
        for row_num, row in enumerate(queryset, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num+1, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response
    else:
        return redirect(index)


def download_file(request):
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    file_name = 'inventory_report.xlsx'
    # Define the full file path
    filepath = BASE_DIR + '/home/Files/' + file_name
    file_name = os.path.basename(filepath)
    chunk_size = 8192
    response = StreamingHttpResponse(FileWrapper(open(filepath, 'rb'), chunk_size), 
                                        content_type=mimetypes.guess_type(filepath)[0])
    response['Content-Length'] = os.path.getsize(filepath)
    response['Content-Disposition'] = "Attachment;filename=%s" % file_name
    return response
